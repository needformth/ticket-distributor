import csv
from pathlib import Path
from openpyxl import load_workbook, Workbook

class FileHandler:
    @staticmethod
    def load_students(file_path: str, has_name: bool = True) -> list[str]:
        """
        Загружает список студентов из CSV или Excel (первый столбец)

        Args:
            file_path: путь до файла
            has_name: есть ли строка с заголовком (если True, первая строка пропускается)

        Returns:
            Список студентов

        Raises:
            ValueError: если формат файла не поддерживается
        """
        ext = Path(file_path).suffix.lower()
        if ext not in ('.csv', '.xlsx'):
            raise ValueError(f"Неподдерживаемый формат файла: {ext}. Поддерживаются .csv и .xlsx")

        students = []
        if ext == '.csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                # Пропускаем заголовок, если нужно
                if has_name:
                    next(reader, None)
                for row in reader:
                    if row and row[0].strip():
                        students.append(row[0].strip())
        else:  # .xlsx
            wb = load_workbook(file_path)
            ws = wb.active
            # Определяем, с какой строки начинать
            start_row = 2 if has_name else 1
            for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
                if row[0] and str(row[0]).strip():
                    students.append(str(row[0]).strip())
        return students

    @staticmethod
    def save_results(distribution: dict[str: list[int]], file_path: str) -> None:
        """
        Сохраняет распределение билетов в Excel-файл.

        Args:
            distribution: словарь {студент: [номера_билетов]}
            file_path: путь для сохранения (должен заканчиваться на .xlsx)
            to_file: если True, сохраняет файл (по умолчанию True)

        Raises:
            ValueError: если формат файла не .xlsx или путь не указан
        """
        if not file_path:
            raise ValueError("Не указан путь для сохранения файла")
        if not file_path.endswith('.xlsx'):
            raise ValueError(f"Неподдерживаемый формат файла: {file_path}. Требуется .xlsx")

        # Создаём книгу и лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"

        # Заголовки
        ws.append(["Студент", "Билеты"])

        # Данные
        for student, tickets in distribution.items():
            tickets_str = ", ".join(map(str, tickets)) if tickets else "нет билетов"
            ws.append([student, tickets_str])

        # Сохраняем
        wb.save(file_path)